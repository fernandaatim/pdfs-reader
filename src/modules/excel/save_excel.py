import openpyxl
from openpyxl.styles import NamedStyle

def save_data_to_excel(data, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.append(['Empresa', 'Codigo de Faturamento', 'Valor Bruto'])

    number_style = NamedStyle(name="number_style", number_format="0.00")
    
    for item in data:
        empresa = item['empresa']
        codigo = item['codigo']
        total = item['total']
        
        if isinstance(total, str):
            try:
                total = total.replace('.', '')
                total = float(total.replace(",", "."))
            except ValueError:
                total = None
        
        ws.append([empresa, codigo, total])

    for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.style = number_style

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(path)
